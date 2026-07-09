"""techdirect-register-unsellable: TechDirect営業対象外候補者をFreelanceBaseに自動登録する

使い方:
  # 環境変数を確認
  source ~/.zshrc

  # ドライラン（対象収集・氏名判定のみ）
  python3 register_unsellable.py --dry-run > /tmp/register_unsellable.log 2>&1 &

  # 本番実行（FreelanceBase 登録＋TechDirect ラベル付与まで）
  python3 register_unsellable.py > /tmp/register_unsellable.log 2>&1 &

  # スモークテスト: 先頭N件のみ処理
  python3 register_unsellable.py --dry-run --limit 5
  python3 register_unsellable.py --limit 1  # 本番登録1件だけ

OTP_REQUIRED がログに出たら:
  1. Gmail MCPで `label:"01D.Claude" subject:ワンタイムパスワード newer_than:3m` を検索
  2. echo "{6桁OTP}" > /tmp/techdirect_otp.txt
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import threading
import time
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib import request as urlrequest

sys.path.insert(0, __import__("os").path.expanduser("~/.claude/snippets"))

from playwright.sync_api import Locator, Page, sync_playwright

import playwright_freelancebase as fb
import playwright_techdirect as td
from freelancebase.candidates import candidate_id, find_candidate_matches, search_candidates

# ========== 定数 ==========

# 実行時の書き出し先はスキルの state/ 配下に集約する。
# `.claude/skills/techdirect-register-unsellable/scripts/register_unsellable.py`
# → parent.parent が skill root。そこに state/ を作成する
SKILL_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = SKILL_ROOT / "state"
OUT_DIR.mkdir(parents=True, exist_ok=True)

REGISTERED_LABEL = "Base登録済/営業対象外"

# FreelanceBase のコメントタブに残す定型文テンプレート
COMMENT_TEMPLATE = (
    "TechDirect上で営業対象外、または書類見送り人材のため自動登録\n"
    "{detail_url}"
)

# FreelanceBase
FB_DRAWER_SEL = ".modal_content.right-modal.right-open"
FB_UPDATE_API_RE = re.compile(r"/api/enterprise/candidates/update/(\d+)")
FB_CLAUDE_ACCOUNT_ID = "840"  # 自社担当者1 の Claude アカウント ID
FB_RECRUIT_STATUS_VALUE = "3"  # 集客ステータス: 個別連絡
FB_SALES_STATUS_VALUE = "3"    # 営業ステータス: 営業不可

SOURCES = [
    {
        "key": "applications_rejected",
        "label": "お見送り（書類選考）",
        "url": (
            "https://techdirect.jp/orgs/39336/portal/applications/"
            "?recruitment_status_type=custom&recruitment_status_id=10959"
            "&include_archived_and_muted=true"
        ),
        "mode": "infinite_scroll",
    },
    {
        "key": "scout_list_1302",
        "label": "営業対象外",
        "url": (
            "https://techdirect.jp/orgs/39336/portal/job-seekers-tabular"
            "?filters=%5B%7B%22type%22%3A%22multiple%22,%22name%22%3A%22job_seeker_list_id%22,"
            "%22invert%22%3Afalse,%22value%22%3A%5B%221302%22%5D%7D%5D"
        ),
        "mode": "pagination",
    },
    {
        "key": "scout_list_1602",
        "label": "案件が見つかるか不明",
        "url": (
            "https://techdirect.jp/orgs/39336/portal/job-seekers-tabular"
            "?filters=%5B%7B%22type%22%3A%22multiple%22,%22name%22%3A%22job_seeker_list_id%22,"
            "%22invert%22%3Afalse,%22value%22%3A%5B%221602%22%5D%7D%5D"
        ),
        "mode": "pagination",
    },
]

SLACK_WEBHOOK_URL = os.environ.get("SLACK_WEBHOOK_NOTIFICATION_URL")

# セッションキャッシュ（OTP頻度を抑えるため）
TD_STATE_PATH = Path("/tmp/td_state.json")
FB_STATE_PATH = Path("/tmp/fb_state.json")

# 処理済みURLキャッシュ（定期実行時の差分処理用）
# 終状態（valid+label_applied / skip_*）をキャッシュし、2回目以降は
# TD詳細ページアクセスをスキップする
PROCESSED_CACHE_PATH = OUT_DIR / "processed_urls.json"
_TERMINAL_SKIP_STATUSES = {
    "skip_label", "skip_private", "skip_partial", "skip_initial", "skip_nickname",
    # FB既存同姓同名で登録スキップしたもの。再実行時も同じ結果になるので終状態扱い
    "skip_duplicate",
}


# ========== データクラス ==========

@dataclass
class CandidateRef:
    source_key: str
    source_label: str
    detail_url: str  # TechDirect候補者詳細URL
    display_name: str  # 一覧で見えた表示名（nickname/mac等）


@dataclass
class Judgment:
    ref: CandidateRef
    status: str  # "valid" | "skip_label" | "skip_private" | "skip_partial" | "skip_initial" | "skip_nickname" | "error"
    last_name: str | None = None
    first_name: str | None = None
    reason: str = ""
    error: str | None = None


@dataclass
class RegistrationResult:
    judgment: Judgment
    registered: bool = False
    fb_internal_id: int | None = None
    fb_email: str | None = None
    label_applied: bool = False
    error: str | None = None
    skipped_duplicate: bool = False
    duplicate_ids: list[int] | None = None


# ========== ユーティリティ ==========

def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def load_processed_cache() -> dict[str, dict]:
    """processed_urls.json を読み込む。ファイル欠落・破損時は空dict。"""
    if not PROCESSED_CACHE_PATH.exists():
        return {}
    try:
        data = json.loads(PROCESSED_CACHE_PATH.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return data
        return {}
    except Exception as exc:
        log(f"[WARN] processed_urls.json の読込失敗: {exc}")
        return {}


def save_processed_cache(cache: dict[str, dict]) -> None:
    """processed_urls.json を書き出す。ディレクトリは起動時に作成済み。"""
    try:
        PROCESSED_CACHE_PATH.write_text(
            json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )
    except Exception as exc:
        log(f"[WARN] processed_urls.json の書込失敗: {exc}")


def is_terminal_cache_entry(entry: dict) -> bool:
    """キャッシュエントリが「詳細アクセス不要な終状態」か判定.

    終状態:
      - status が skip_* 系（skip_label / skip_private / skip_partial / skip_initial / skip_nickname）
      - status=valid かつ fb_internal_id あり かつ label_applied=True
    非終状態:
      - status=error
      - valid だが label_applied=False（登録未完）
      - status未設定
    """
    status = entry.get("status")
    if not status:
        return False
    if status in _TERMINAL_SKIP_STATUSES:
        return True
    if status == "valid":
        return bool(entry.get("fb_internal_id")) and bool(entry.get("label_applied"))
    return False


def notify_slack(text: str) -> None:
    if not SLACK_WEBHOOK_URL:
        log("[WARN] SLACK_WEBHOOK_NOTIFICATION_URL 未設定。Slack通知スキップ")
        return
    try:
        body = json.dumps({"text": text}).encode("utf-8")
        req = urlrequest.Request(
            SLACK_WEBHOOK_URL,
            data=body,
            headers={"Content-Type": "application/json"},
        )
        with urlrequest.urlopen(req, timeout=10) as resp:
            resp.read()
        log("[INFO] Slack通知送信完了")
    except Exception as exc:
        log(f"[WARN] Slack通知失敗: {exc}")


# ========== セッションキャッシュ付きログイン ==========

def td_login_with_cache(pw, headless: bool = True):
    """TechDirect にログイン。セッションキャッシュがあれば再利用して OTP をスキップする."""
    if TD_STATE_PATH.exists():
        log(f"[INFO] TechDirectセッションキャッシュ利用: {TD_STATE_PATH}")
        browser = pw.chromium.launch(headless=headless)
        context = browser.new_context(storage_state=str(TD_STATE_PATH))
        page = context.new_page()
        # セッション有効性を確認（/jobs 等の保護ページにアクセスして判定）
        page.goto("https://techdirect.jp/jobs", wait_until="networkidle")
        page.wait_for_timeout(2000)
        if "login" not in page.url:
            log(f"[INFO] セッション有効 URL={page.url}")
            return browser, context, page
        log("[INFO] セッションキャッシュ無効。ログインフォールバック")
        browser.close()

    browser, context, page = td.login(pw, headless=headless)
    # セッションを保存
    try:
        context.storage_state(path=str(TD_STATE_PATH))
        log(f"[INFO] TechDirectセッション保存: {TD_STATE_PATH}")
    except Exception as exc:
        log(f"[WARN] セッション保存失敗: {exc}")
    return browser, context, page


def fb_login_with_cache(pw, headless: bool = True):
    """FreelanceBase にログイン。セッションキャッシュがあれば再利用."""
    if FB_STATE_PATH.exists():
        log(f"[INFO] FreelanceBaseセッションキャッシュ利用: {FB_STATE_PATH}")
        browser = pw.chromium.launch(headless=headless)
        # UA必須
        context = browser.new_context(
            storage_state=str(FB_STATE_PATH),
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()
        page.goto(
            "https://freelancebase.jp/enterprise/candidates",
            wait_until="networkidle",
        )
        page.wait_for_timeout(2000)
        if "signin" not in page.url:
            log(f"[INFO] FBセッション有効 URL={page.url}")
            return browser, context, page
        log("[INFO] FBセッションキャッシュ無効。ログインフォールバック")
        browser.close()

    browser, context, page = fb.login(pw, headless=headless)
    try:
        context.storage_state(path=str(FB_STATE_PATH))
        log(f"[INFO] FBセッション保存: {FB_STATE_PATH}")
    except Exception as exc:
        log(f"[WARN] FBセッション保存失敗: {exc}")
    return browser, context, page


# ========== 並列判定用ワーカー ==========

# 各ワーカースレッドが独立した Playwright / ブラウザ / コンテキスト / ページを持つ
_worker_local = threading.local()
# 終了時に cleanup するためのワーカー一覧（pw, browser のペア）
_worker_instances: list[tuple] = []
_worker_instances_lock = threading.Lock()


def _get_worker_page() -> Page:
    """スレッドごとに独立した Playwright ページを返す（storage_state 共有でログイン済み）"""
    page = getattr(_worker_local, "page", None)
    if page is not None:
        return page
    from playwright.sync_api import sync_playwright as _spw

    pw = _spw().start()
    browser = pw.chromium.launch(headless=True)
    context = browser.new_context(storage_state=str(TD_STATE_PATH))
    context.set_default_timeout(45000)
    page = context.new_page()
    _worker_local.pw = pw
    _worker_local.browser = browser
    _worker_local.context = context
    _worker_local.page = page
    with _worker_instances_lock:
        _worker_instances.append((pw, browser))
    return page


def _cleanup_workers() -> None:
    """ワーカー由来の Playwright リソースを全て閉じる"""
    with _worker_instances_lock:
        snapshot = list(_worker_instances)
        _worker_instances.clear()
    for pw, browser in snapshot:
        try:
            browser.close()
        except Exception:
            pass
        try:
            pw.stop()
        except Exception:
            pass


# ========== 並列登録用ワーカー（FB + TD のペアをスレッドごとに保持） ==========

_reg_worker_local = threading.local()
_reg_worker_instances: list[tuple] = []
_reg_worker_instances_lock = threading.Lock()


def _get_worker_reg_pages() -> tuple[Page, Page]:
    """スレッドごとに FB ページ / TD ページの独立ペアを返す (fb_page, td_page).

    登録フェーズを並列化するため、各スレッドが独立した Chromium コンテキストを持つ.
    storage_state で既存ログインを共有するので OTP / 再ログインは不要.
    """
    if getattr(_reg_worker_local, "initialized", False):
        return _reg_worker_local.fb_page, _reg_worker_local.td_page
    from playwright.sync_api import sync_playwright as _spw

    pw = _spw().start()

    td_browser = pw.chromium.launch(headless=True)
    td_context = td_browser.new_context(storage_state=str(TD_STATE_PATH))
    td_context.set_default_timeout(45000)
    td_page = td_context.new_page()

    fb_browser = pw.chromium.launch(headless=True)
    fb_context = fb_browser.new_context(
        storage_state=str(FB_STATE_PATH),
        user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
        ),
    )
    fb_context.set_default_timeout(45000)
    fb_page = fb_context.new_page()

    _reg_worker_local.pw = pw
    _reg_worker_local.td_browser = td_browser
    _reg_worker_local.td_page = td_page
    _reg_worker_local.fb_browser = fb_browser
    _reg_worker_local.fb_page = fb_page
    _reg_worker_local.initialized = True
    with _reg_worker_instances_lock:
        _reg_worker_instances.append((pw, td_browser, fb_browser))
    return fb_page, td_page


def _cleanup_reg_workers() -> None:
    with _reg_worker_instances_lock:
        snapshot = list(_reg_worker_instances)
        _reg_worker_instances.clear()
    for pw, td_browser, fb_browser in snapshot:
        try:
            td_browser.close()
        except Exception:
            pass
        try:
            fb_browser.close()
        except Exception:
            pass
        try:
            pw.stop()
        except Exception:
            pass


# ========== 氏名判定 ==========

# 英字のみ(大文字小文字ラテン + 数字) & 1〜2語構成 & 姓/名どちらか不明
# → nickname または partial
_JAP_NAME_RE = re.compile(r"^[一-龯ぁ-んァ-ヶー々]+$")
_LATIN_ONLY_RE = re.compile(r"^[A-Za-z0-9._\- ]+$")


def classify_name(raw: str) -> tuple[str, str | None, str | None, str]:
    """氏名文字列を分類する.

    Returns:
        (status, last_name, first_name, reason)
        status は "valid" / "skip_private" / "skip_partial" / "skip_initial" / "skip_nickname" のいずれか
    """
    if raw is None:
        return "skip_private", None, None, "氏名 None"
    name = raw.strip()
    if not name or "非公開" in name or name == "未設定":
        return "skip_private", None, None, f"非公開/未設定: '{name}'"

    # イニシャル判定: . を含む、または2文字以下の大文字主体
    if "." in name:
        return "skip_initial", None, None, f"イニシャル含み: '{name}'"
    if len(name) <= 2 and _LATIN_ONLY_RE.match(name):
        return "skip_initial", None, None, f"2文字以下のラテン: '{name}'"

    # 姓名分割: 全角/半角スペース or 中黒
    tokens = [t for t in re.split(r"[\s　・]+", name) if t]
    if len(tokens) < 2:
        # トークンが1つしかない場合
        # 日本語名で姓＋名がスペースなく結合されているケースもあるが、
        # その場合は判定不能なのでスキップ（ユーザー要求に従い「姓のみ/名のみは除外」）
        if _LATIN_ONLY_RE.match(name):
            return "skip_nickname", None, None, f"単一トークン（ラテン/英字のみ → ハンドル名可能性）: '{name}'"
        return "skip_partial", None, None, f"姓名分割不可（スペース等で区切れない）: '{name}'"

    # 2トークン以上: 先頭を姓、2つ目を名とする
    last_name, first_name = tokens[0], tokens[1]

    # 各トークンが1文字のみ（例: "A B" "山 田"）はイニシャル扱い
    if len(last_name) == 1 or len(first_name) == 1:
        return "skip_initial", None, None, f"1文字トークン: '{last_name}' / '{first_name}'"

    # ラテンのみで構成されていて、本当に本名かどうか疑わしい場合は nickname 扱い
    # （ただし "John Smith" 等も除外されるのでここは慎重に）
    # → 今回は要求通り「明らかなあだ名」のみ除外。ラテン姓名は valid とする
    if _LATIN_ONLY_RE.match(name):
        # アンダースコア、数字を含む場合は nickname とみなす
        if re.search(r"[0-9_]", name):
            return "skip_nickname", None, None, f"数字/アンダースコア含み: '{name}'"

    return "valid", last_name, first_name, f"姓名: '{last_name}' / '{first_name}'"


# ========== TechDirect 収集ロジック ==========

def collect_from_applications(page: Page, url: str) -> list[CandidateRef]:
    """「お見送り（書類選考）」メッセージ一覧から収集.

    TDのapplicationsページはVue Router + 内部スクロールコンテナ構造で、
    `document.body` のスクロール・`<a href>` セレクタは機能しない。
    代わりに内部スクロールコンテナを scrollTop 操作し、
    `api.codeal.work/v1/orgs/39336/applications` のXHRレスポンスを
    `page.on("response")` で捕捉して候補者UUIDを抽出する。
    """
    log(f"[STEP] 収集開始（内部スクロール+API応答捕捉）: {url}")

    captured_users: dict[str, dict] = {}

    def _on_resp(resp):
        # お見送り一覧の API のみに限定。他の関連 API（unread_ids/is_high_priority等）は除外
        if "api.codeal.work/v1/orgs/" not in resp.url or "/applications" not in resp.url:
            return
        if "unread" in resp.url or "is_high_priority" in resp.url or "limit=1&" in resp.url or resp.url.endswith("limit=1"):
            return
        try:
            if resp.status != 200:
                return
            body = resp.text()
            data = json.loads(body)
        except Exception:
            return
        # Codeal APIは配列（list）でレスポンスを返す仕様
        if isinstance(data, list):
            apps = data
        elif isinstance(data, dict):
            apps = data.get("applications") or []
        else:
            apps = []
        for app in apps:
            user = app.get("user") or {}
            # url_key が UUID に相当する（TDの候補者詳細URLは /users/{url_key}）
            uuid = user.get("url_key") or user.get("uuid") or app.get("user_url_key")
            # 氏名候補: fullname → name → display_name
            name = user.get("fullname") or user.get("name") or user.get("display_name") or ""
            if uuid:
                captured_users[uuid] = {"uuid": uuid, "name": str(name).strip()}

    page.on("response", _on_resp)
    try:
        page.goto(url, wait_until="networkidle")
        page.wait_for_timeout(3500)

        # 内部スクロール: overflow:auto/scroll の最大 scrollHeight コンテナを
        # 繰り返し最下部へスクロールして追加ロードを誘発する
        max_iters = 120  # 699件想定 × limit=10 → 70+α
        stale_streak = 0
        prev_count = 0
        for i in range(max_iters):
            page.evaluate(
                """() => {
                    let target = null, maxH = 0;
                    document.querySelectorAll('*').forEach(el => {
                        const s = getComputedStyle(el);
                        if ((s.overflowY === 'auto' || s.overflowY === 'scroll')
                            && el.scrollHeight > el.clientHeight
                            && el.scrollHeight > maxH) {
                            maxH = el.scrollHeight;
                            target = el;
                        }
                    });
                    if (target) target.scrollTop = target.scrollHeight;
                }"""
            )
            page.wait_for_timeout(1800)
            cur = len(captured_users)
            if cur == prev_count:
                stale_streak += 1
                if stale_streak >= 3:
                    log(f"[INFO] 追加ロード停止 iter={i+1} total={cur}")
                    break
            else:
                stale_streak = 0
            prev_count = cur
    finally:
        try:
            page.remove_listener("response", _on_resp)
        except Exception:
            pass

    refs: list[CandidateRef] = []
    for uuid, info in captured_users.items():
        detail_url = f"https://techdirect.jp/users/{uuid}?from_org_id=39336"
        refs.append(
            CandidateRef(
                source_key="applications_rejected",
                source_label="お見送り（書類選考）",
                detail_url=detail_url,
                display_name=info.get("name") or "",
            )
        )
    log(f"[INFO] applications: {len(refs)} 件")
    return refs


def collect_from_scout_list(
    page: Page,
    url: str,
    source_key: str,
    source_label: str,
    collect_cap: int | None = None,
) -> list[CandidateRef]:
    """「営業対象外」「案件が見つかるか不明」のテーブル形式（ページネーション）から収集.

    collect_cap を指定すると件数が上限に達した時点で収集を打ち切る.
    """
    log(f"[STEP] 収集開始（ページネーション）: {source_label}")
    page.goto(url, wait_until="networkidle")
    page.wait_for_timeout(3000)

    refs: list[CandidateRef] = []
    seen: set[str] = set()
    page_no = 1
    max_pages = 60  # 安全弁
    zero_new_streak = 0

    while page_no <= max_pages:
        page.wait_for_timeout(2000)
        handles = page.locator("a[href*='/users/']").element_handles()
        new_count = 0
        for h in handles:
            href = h.get_attribute("href") or ""
            txt = (h.inner_text() or "").strip()
            full_url = href if href.startswith("http") else f"https://techdirect.jp{href}"
            m = re.search(r"/users/([0-9a-f-]{36})", full_url)
            uid = m.group(1) if m else full_url
            if uid in seen:
                continue
            seen.add(uid)
            new_count += 1
            refs.append(
                CandidateRef(
                    source_key=source_key,
                    source_label=source_label,
                    detail_url=full_url,
                    display_name=txt,
                )
            )
        log(f"[INFO] {source_label} page {page_no}: 新規 {new_count} 件 / 累計 {len(refs)} 件")

        # 0 件が 2 ページ続いたら終端とみなす
        if new_count == 0:
            zero_new_streak += 1
            if zero_new_streak >= 2:
                log(f"[INFO] {source_label}: 新規0件 x{zero_new_streak} のため終了")
                break
        else:
            zero_new_streak = 0

        # collect_cap の早期打ち切り
        if collect_cap is not None and len(refs) >= collect_cap:
            log(f"[INFO] {source_label}: collect_cap={collect_cap} に到達、収集終了")
            break

        # 「次へ」ボタンを探す
        next_btn = page.locator(
            "button:has-text('次へ'), button[aria-label='Next'], a:has-text('次へ')"
        ).first
        if next_btn.count() == 0 or not next_btn.is_enabled():
            log(f"[INFO] {source_label}: 次ページボタン未検出/非活性のため終了")
            break
        try:
            next_btn.click()
        except Exception as exc:
            log(f"[WARN] 次ページボタンクリック失敗: {exc}")
            break
        page.wait_for_timeout(2500)
        page_no += 1

    log(f"[INFO] {source_label}: 全 {len(refs)} 件 ({page_no} ページ)")
    return refs


# ========== 候補者詳細から判定 ==========

def judge_candidate(page: Page, ref: CandidateRef) -> Judgment:
    """候補者詳細ページを開き、ラベル存在＆氏名判定を行う.

    氏名フィールドが 非公開 / 部分的な場合は、添付スキルシート・職務経歴書から
    氏名抽出を試みるフォールバックを含む。

    DNS / ネットワーク一時障害に備えて page.goto を最大3回リトライする
    （バックオフ: 3秒 → 8秒）.
    """
    last_exc: Exception | None = None
    for attempt in range(3):
        try:
            page.goto(ref.detail_url, wait_until="networkidle")
            page.wait_for_timeout(2000)
            last_exc = None
            break
        except Exception as exc:
            last_exc = exc
            if attempt < 2:
                backoff = 3 + attempt * 5
                log(f"[WARN] 詳細遷移失敗 attempt={attempt+1}/3 ({exc}) → {backoff}s 待機")
                time.sleep(backoff)
    if last_exc is not None:
        return Judgment(ref=ref, status="error", error=f"詳細ページ遷移失敗 (3回リトライ): {last_exc}")

    # ラベル判定
    already_labeled = check_registered_label(page)
    if already_labeled:
        return Judgment(ref=ref, status="skip_label", reason="既にBase登録済みラベルあり")

    # 氏名取得: 詳細画面の「氏名」セル
    name_value = extract_name_on_detail(page)
    status, last, first, reason = classify_name(name_value or ref.display_name)

    # skip_private / skip_partial の場合は添付ファイルからのフォールバック抽出を試みる
    if status in ("skip_private", "skip_partial"):
        att_pair = extract_name_from_attachments(page)
        if att_pair:
            att_last, att_first = att_pair
            status2, last2, first2, reason2 = classify_name(f"{att_last} {att_first}")
            if status2 == "valid":
                reason = f"添付から取得: '{att_last} {att_first}' (元氏名field: {reason})"
                return Judgment(
                    ref=ref,
                    status="valid",
                    last_name=last2,
                    first_name=first2,
                    reason=reason,
                )

    return Judgment(
        ref=ref,
        status=status,
        last_name=last,
        first_name=first,
        reason=reason,
    )


# ========== 添付ファイルからの氏名抽出 ==========

# 履歴書・スキルシート中によく現れるフォーム項目ラベル（氏名の次トークンとして間違えないよう除外）
_LABEL_BLACKLIST = {
    # 分割可能な2トークンフォームラベル（各パーツ単位も追加）
    "生年月日", "生年", "月日",
    "最終学歴", "最終",
    "保有資格", "保有",
    "職務経歴", "職務", "職歴", "経歴", "職歴書",
    "職務経歴書", "履歴書", "スキルシート",
    # 履歴書/職務経歴書のセクション見出し（氏名直後に来て誤マッチしやすい）
    "職務要約", "職務概要", "経歴概要", "経歴要約", "基本情報", "個人情報",
    "資格免許", "免許資格", "要約", "概要",
    "フリガナ", "ふりがな", "フリ", "ガナ",
    # よくあるフォーム項目
    "年齢", "性別", "住所", "電話", "電話番号", "連絡先", "メール", "メールアドレス",
    "E-mail", "Email", "学歴", "資格", "スキル", "免許",
    "国籍", "就業", "就業可能時期", "稼働", "稼働開始", "最寄", "最寄駅",
    "セイ", "メイ", "姓", "名", "名前", "氏名", "職種", "希望", "備考", "自己PR",
    "出身", "出身地", "居住地", "現住所", "FAX", "LINE",
    # 期間 / スケジュール系
    "期間", "開始", "終了", "就業時期", "期限",
    # 学校 / 教育系
    "大学", "大学院", "高校", "高等学校", "専門学校", "短大", "卒業",
    # プロジェクト / 業務系（履歴書内のヘッダ行）
    "プロジェクト", "業務", "業務内容", "役職", "ポジション",
    # 自己PR / 備考
    "自己紹介", "備考欄", "PR",
}


def _is_valid_name_token(token: str) -> bool:
    """氏名トークンとして妥当か（ラベル語でなく・2〜6文字）を判定"""
    if not token or token in _LABEL_BLACKLIST:
        return False
    if not (2 <= len(token) <= 6):
        return False
    return True


def _split_concat_name(concat: str) -> tuple[str, str] | None:
    """連結された姓名（4-6文字）を妥当な (姓, 名) に分割する。失敗時 None。

    優先順位:
      4字: (2+2)
      5字: (3+2) → (2+3)  ※ 長谷川由美, 田中井夕侑 のような 3字姓を優先
      6字: (3+3) → (2+4)
    """
    n = len(concat)
    if n == 4:
        splits = [2]
    elif n in (5, 6):
        splits = [3, 2]
    else:
        splits = [2, 3]
    for split in splits:
        if split >= n:
            continue
        last, first = concat[:split], concat[split:]
        if _is_valid_name_token(last) and _is_valid_name_token(first) and last != first:
            return last, first
    return None


# 最初のパス: ラベル付き（「氏名: 山田 太郎」「氏名 山田 太郎」）
_ATT_NAME_PATTERN_LABELED = re.compile(
    r"(?:氏\s*名|名\s*前|Name)\s*[:：\s]\s*([一-龯ぁ-んァ-ヶー]{2,6})\s+([一-龯ぁ-んァ-ヶー]{2,6})"
)
# 2番目のパス: ラベル付きで姓名が連結された場合（「氏名: 山田太郎」4-6文字）
_ATT_NAME_PATTERN_LABELED_CONCAT = re.compile(
    r"(?:氏\s*名|名\s*前|Name)\s*[:：\s]\s*([一-龯ぁ-んァ-ヶー]{4,6})\s"
)


def extract_name_from_attachments(page: Page) -> tuple[str, str] | None:
    """TD候補者詳細から添付ファイル（スキルシート/職務経歴書）をDLし氏名抽出を試みる.

    対応拡張子: .pdf / .xlsx / .docx（.xls は非対応）
    戻り値: (姓, 名) or None
    """
    import tempfile

    try:
        attachments = page.evaluate(
            """() => {
                const result = [];
                const anchors = document.querySelectorAll('a[href^="/media/portfolio-files/"]');
                for (const a of anchors) {
                    result.push({
                        filename: (a.innerText || '').trim(),
                        href: a.getAttribute('href'),
                    });
                }
                return result;
            }"""
        )
    except Exception as exc:
        log(f"[WARN] 添付リスト取得失敗: {exc}")
        return None

    if not attachments:
        return None

    api = page.context.request
    for att in attachments:
        filename = att.get("filename", "")
        href = att.get("href", "")
        ext = Path(filename).suffix.lower()
        if ext not in (".pdf", ".xlsx", ".docx"):
            log(f"[INFO] 添付 '{filename}' 拡張子 {ext} は抽出未対応")
            continue
        url = "https://techdirect.jp" + href
        try:
            resp = api.get(url, timeout=15000)
            if resp.status != 200:
                log(f"[WARN] 添付DL失敗 status={resp.status}")
                continue
            data = resp.body()
        except Exception as exc:
            log(f"[WARN] 添付DL例外: {exc}")
            continue

        text = ""
        try:
            with tempfile.NamedTemporaryFile(suffix=ext, delete=True) as tmp:
                tmp.write(data)
                tmp.flush()
                if ext == ".pdf":
                    try:
                        import pdfplumber  # type: ignore
                    except ImportError:
                        log("[WARN] pdfplumber 未インストール")
                        continue
                    with pdfplumber.open(tmp.name) as pdf:
                        for p in pdf.pages[:3]:
                            text += (p.extract_text() or "") + "\n"
                elif ext == ".xlsx":
                    try:
                        from openpyxl import load_workbook  # type: ignore
                    except ImportError:
                        log("[WARN] openpyxl 未インストール")
                        continue
                    wb = load_workbook(tmp.name, read_only=True, data_only=True)
                    for sh in wb.worksheets[:1]:
                        for row in sh.iter_rows(max_row=30, values_only=True):
                            for cell in row:
                                if cell:
                                    text += str(cell) + " "
                            text += "\n"
                    wb.close()
                elif ext == ".docx":
                    try:
                        import docx  # type: ignore
                    except ImportError:
                        log("[WARN] python-docx 未インストール")
                        continue
                    d = docx.Document(tmp.name)
                    for p in d.paragraphs[:30]:
                        text += p.text + "\n"
        except Exception as exc:
            log(f"[WARN] 添付テキスト抽出失敗 ({filename}): {exc}")
            continue

        if not text.strip():
            continue

        head = text[:2000]

        # パターンA: ラベル付き + スペース区切り 姓 名
        m = _ATT_NAME_PATTERN_LABELED.search(head)
        if m:
            last, first = m.group(1).strip(), m.group(2).strip()
            if _is_valid_name_token(last) and _is_valid_name_token(first) and last != first:
                log(f"[INFO] 添付 '{filename}' から氏名抽出成功: {last} {first}")
                return last, first
            # group2 が見出し語など妥当でない場合: 「氏名 桑原宗梧 職務要約」のように
            # group1 が連結フルネーム、group2 がセクション見出しのケース。
            # group1 を連結名とみなして分割を試みる。
            if _is_valid_name_token(last) and 4 <= len(last) <= 6:
                split = _split_concat_name(last)
                if split:
                    log(
                        f"[INFO] 添付 '{filename}' から氏名抽出（パターンA group2無効→"
                        f"group1連結分割）: {split[0]} / {split[1]} (元: {last}, "
                        f"無効group2: '{first}')"
                    )
                    return split
            log(
                f"[DEBUG] 添付 '{filename}' パターンA マッチしたが妥当でない: "
                f"'{last}' / '{first}'"
            )

        # パターンB: ラベル付き + 姓名連結（4-6文字）→ 2+2 / 3+2 / 2+3 / 3+3 で分割
        m = _ATT_NAME_PATTERN_LABELED_CONCAT.search(head)
        if m:
            concat = m.group(1).strip()
            split = _split_concat_name(concat)
            if split:
                log(
                    f"[INFO] 添付 '{filename}' から氏名抽出（連結分割）: "
                    f"{split[0]} / {split[1]} (元: {concat})"
                )
                return split

    return None


def check_registered_label(page: Page) -> bool:
    """候補者詳細ページでBase登録済みラベルの付与有無を判定.

    「リスト」ボタン → ドロップダウン内の
      <div class="custom-control custom-checkbox">
        <input type="checkbox" id="__cid__N">
        <label for="__cid__N"><div>Base登録済み（営業対象外）</div></label>
      </div>
    構造を辿り、input.checked を見る。
    """
    try:
        btn = page.locator("button:has-text('リスト')").first
        if btn.count() == 0:
            return False
        btn.click()
        page.wait_for_timeout(800)

        # JS で label→input を辿って checked 判定
        checked = page.evaluate(
            """(targetText) => {
                const labels = Array.from(document.querySelectorAll('label.custom-control-label'));
                for (const lb of labels) {
                    const text = (lb.innerText || '').trim();
                    if (text === targetText) {
                        const id = lb.getAttribute('for');
                        if (!id) continue;
                        const input = document.getElementById(id);
                        if (input) {
                            return { found: true, checked: !!input.checked };
                        }
                    }
                }
                return { found: false, checked: false };
            }""",
            REGISTERED_LABEL,
        )

        # ドロップダウンを閉じる
        page.keyboard.press("Escape")
        page.wait_for_timeout(500)

        if not checked.get("found"):
            log(f"[WARN] ドロップダウンに '{REGISTERED_LABEL}' label 未検出")
            return False
        return bool(checked.get("checked"))
    except Exception as exc:
        log(f"[WARN] ラベル判定エラー ({page.url}): {exc}")
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return False


def extract_name_on_detail(page: Page) -> str | None:
    """詳細画面から氏名フィールドの値を取得する。非公開なら '非公開' を返す。

    DOM: `<dt><span>氏名</span></dt><dd>値</dd>` 形式（dt/dd）
    """
    try:
        result = page.evaluate(
            """() => {
                const dts = Array.from(document.querySelectorAll('dt'));
                for (const dt of dts) {
                    const label = (dt.innerText || '').replace(/\\s+/g, '').trim();
                    if (label === '氏名') {
                        let sib = dt.nextElementSibling;
                        while (sib && sib.tagName !== 'DD') sib = sib.nextElementSibling;
                        if (sib) return (sib.innerText || '').trim();
                    }
                }
                return null;
            }"""
        )
        return result or None
    except Exception as exc:
        log(f"[WARN] 氏名抽出エラー: {exc}")
        return None


# ========== FreelanceBase 登録ロジック ==========

def check_fb_duplicate(fb_page: Page, last_name: str, first_name: str) -> list[int]:
    """FreelanceBase で同姓同名の既存候補者を探す（API直叩き版）.

    旧実装の「クイック検索」入力はテーブルをフィルタしないため、デフォルト50行の中に
    検索対象がいないと見逃していた。本実装は `/api/enterprise/candidates/index` に
    `keyword` を渡して正確に検索する。

    戻り値: 完全一致する人材の `id_by_enterprise_id` リスト.
    """
    query = f"{last_name} {first_name}"
    candidates = search_candidates(fb_page, query, raise_on_error=True)
    log(f"[INFO] 重複チェック (API) '{query}': {len(candidates)} 件ヒット")

    # 完全一致判定: l_name + f_name 完全一致、または name/supplier_name が空白除去で一致
    matches: list[int] = []
    for candidate in find_candidate_matches(
        candidates,
        last_name=last_name,
        first_name=first_name,
    ):
        cid = candidate_id(candidate)
        if cid is not None and cid not in matches:
            matches.append(cid)
    return matches


def generate_email() -> str:
    return (
        f"sample.claude{datetime.now().strftime('%Y%m%d%H%M%S%f')[:-3]}"
        f"_{uuid.uuid4().hex[:6]}@example.com"
    )


def create_fb_candidate_with_retry(
    fb_page: Page, last_name: str, first_name: str, max_retries: int = 3
) -> tuple[int, str]:
    """create_fb_candidate を 422/タイムアウト等の一時エラーに対してリトライする.

    各試行で異なるメールアドレスを生成（生成済みの email は捨てられる想定）.
    戻り値: (internal_id, 最終的に使用した email)
    """
    last_exc: Exception | None = None
    for attempt in range(max_retries):
        email = generate_email()
        try:
            internal_id = create_fb_candidate(fb_page, last_name, first_name, email)
            return internal_id, email
        except Exception as exc:
            last_exc = exc
            msg = str(exc)
            retriable = "422" in msg or "Timeout" in msg or "timeout" in msg
            if not retriable or attempt == max_retries - 1:
                break
            backoff = 4 + attempt * 6  # 4, 10, 16 秒
            log(
                f"[WARN] create_fb_candidate 失敗 attempt={attempt+1}/{max_retries} "
                f"({msg[:120]}) → {backoff}s 待機して再試行"
            )
            # 残骸ドロワーを閉じる
            try:
                _fb_close_drawer(fb_page)
            except Exception:
                pass
            try:
                fb_page.keyboard.press("Escape")
            except Exception:
                pass
            time.sleep(backoff)
    assert last_exc is not None
    raise last_exc


def create_fb_candidate(
    fb_page: Page, last_name: str, first_name: str, email: str
) -> int:
    """FreelanceBase の「人材を作成」→「通常作成」モーダルで新規候補者を作成し、
    作成後の詳細ページ URL から internal_id を取得して返す."""
    # 人材一覧へ移動（既にいる場合も safe）
    if "/enterprise/candidates" not in fb_page.url:
        fb_page.goto(
            "https://freelancebase.jp/enterprise/candidates",
            wait_until="networkidle",
        )
        fb_page.wait_for_timeout(2000)

    # ドロップダウンを開く
    trigger = fb_page.locator("button:has-text('人材を作成')").first
    if trigger.count() == 0:
        raise RuntimeError("「人材を作成」ボタン未検出")
    trigger.click()
    fb_page.wait_for_timeout(1000)

    # 「通常作成」メニューをクリック
    normal_create = fb_page.locator("text=通常作成").first
    if normal_create.count() == 0:
        raise RuntimeError("ドロップダウン内に「通常作成」メニュー未検出")
    normal_create.click()
    fb_page.wait_for_timeout(2500)

    # モーダルが開いたか確認
    drawer = fb_page.locator(FB_DRAWER_SEL).filter(has_text="人材を作成").first
    if drawer.count() == 0:
        # fallback: 右ドロワーのいずれか
        drawer = fb_page.locator(FB_DRAWER_SEL).first
        if drawer.count() == 0:
            raise RuntimeError("新規作成ドロワー未検出")

    # 所属=自社 (デフォルトで選択済み。明示的な操作はせず触らない)

    # メールアドレス
    email_input = drawer.locator("input[placeholder='contact@freelancebase.jp']").first
    if email_input.count() == 0:
        raise RuntimeError("メールアドレス入力欄未検出")
    _fb_set_text_input(email_input, email)

    # 氏名（姓 / 名）
    last_input = drawer.locator("input[placeholder='田中']").first
    first_input = drawer.locator("input[placeholder='太郎']").first
    if last_input.count() == 0 or first_input.count() == 0:
        raise RuntimeError("氏名入力欄未検出")
    _fb_set_text_input(last_input, last_name)
    _fb_set_text_input(first_input, first_name)

    # 「人材を作成」ボタン（ドロワー下部のみに絞る）
    submit_btn = drawer.locator("button.bg-primary:has-text('人材を作成')").first
    if submit_btn.count() == 0:
        # fallback: type=submit
        submit_btn = drawer.locator("button[type='submit']:has-text('人材を作成')").first
    if submit_btn.count() == 0:
        # 最終 fallback: drawer内の全button中 "人材を作成" で活性なもの
        submit_btn = drawer.locator("button:has-text('人材を作成')").last
    if submit_btn.count() == 0:
        raise RuntimeError("作成送信ボタン未検出")

    # create API 応答から internal_id を取得
    with fb_page.expect_response(
        lambda r: (
            "/api/enterprise/candidates/create" in r.url
            and r.request.method == "POST"
        ),
        timeout=15000,
    ) as resp_info:
        submit_btn.scroll_into_view_if_needed()
        submit_btn.click(force=True)
    resp = resp_info.value
    log(f"[INFO] create API {resp.status} ({resp.request.method} {resp.url})")
    if resp.status >= 400:
        _save_debug_capture(fb_page, "create_fail")
        # 422 はバリデーション/タイミング系の一時エラーの可能性があるため
        # 呼び出し側でリトライ判定できるよう status コード付きエラーを投げる
        raise RuntimeError(f"create API status={resp.status}")

    # レスポンスボディから id を取り出す
    try:
        body = resp.json()
    except Exception as exc:
        _save_debug_capture(fb_page, "create_no_json")
        raise RuntimeError(f"create API レスポンス JSON パース失敗: {exc}")

    log(f"[INFO] create response body (head): {json.dumps(body, ensure_ascii=False)[:500]}")

    # FB の「人材ID」は id_by_enterprise_id。top-level `id` は auth 用 ID（URLでも開けるが中身空）
    internal_id = None
    if isinstance(body, dict):
        for key in ("id_by_enterprise_id", "candidate_id", "internal_id"):
            v = body.get(key)
            if isinstance(v, int):
                internal_id = v
                break

    if internal_id is None:
        log(f"[WARN] APIレスポンスに id を発見できず。body keys={list(body.keys()) if isinstance(body, dict) else type(body)}")
        _save_debug_capture(fb_page, "create_no_id")
        raise RuntimeError("create API レスポンスに internal_id 未検出")

    log(f"[INFO] 新規登録成功 internal_id={internal_id}")

    # モーダルが閉じるのを待ち、詳細ページへ明示的に遷移
    fb_page.wait_for_timeout(2000)
    fb_page.goto(
        f"https://freelancebase.jp/enterprise/candidates/{internal_id}",
        wait_until="networkidle",
    )
    fb_page.wait_for_timeout(2500)
    return internal_id


def _fb_set_text_input(locator: Locator, value: str) -> None:
    """Vue v-model 対応の安全な text 入力. fill() + JS event 発火."""
    locator.scroll_into_view_if_needed()
    locator.click()
    locator.fill("")
    locator.evaluate(
        """(el, v) => {
            el.focus();
            el.value = v;
            el.dispatchEvent(new Event('input', {bubbles:true}));
            el.dispatchEvent(new Event('change', {bubbles:true}));
            el.dispatchEvent(new Event('blur', {bubbles:true}));
        }""",
        value,
    )


def _save_debug_capture(page: Page, prefix: str) -> None:
    """エラー時のデバッグスナップショット保存."""
    try:
        debug_dir = OUT_DIR / "debug"
        debug_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        png = debug_dir / f"{prefix}_{ts}.png"
        html = debug_dir / f"{prefix}_{ts}.html"
        page.screenshot(path=str(png), full_page=True)
        html.write_text(page.content(), encoding="utf-8")
        log(f"[DEBUG] 保存: {png} / {html}")
    except Exception as exc:
        log(f"[WARN] デバッグ保存失敗: {exc}")


def _fb_open_section(page: Page, section_heading: str) -> Locator:
    """候補者詳細から {section_heading} 見出し近傍の「編集する」を押してドロワーを返す."""
    heading = page.locator("h6.fs-18").filter(has_text=section_heading).first
    if heading.count() == 0:
        raise RuntimeError(f"セクション見出し '{section_heading}' 未検出")
    row = heading.locator(
        "xpath=ancestor::div[contains(@class,'row') and contains(@class,'align-items-center')][1]"
    )
    btn = row.get_by_role("button", name="編集する").first
    if btn.count() == 0:
        raise RuntimeError(f"『編集する』ボタン未検出 ({section_heading})")
    btn.scroll_into_view_if_needed()
    btn.click(force=True)
    page.wait_for_selector(FB_DRAWER_SEL, timeout=10000)
    # title がセクション名を含むドロワーを探す
    for _ in range(20):
        page.wait_for_timeout(250)
        drawers = page.locator(FB_DRAWER_SEL)
        for i in range(drawers.count()):
            d = drawers.nth(i)
            try:
                info = d.evaluate(
                    """el => {
                        let title = '';
                        for (const n of el.querySelectorAll('*')) {
                            const t=(n.innerText||'').trim();
                            if (t.endsWith('を編集') && t.length < 40) { title = t; break; }
                        }
                        return { title, inputs: el.querySelectorAll('input, textarea, select').length };
                    }"""
                )
            except Exception:
                continue
            if section_heading in info.get("title", "") and info.get("inputs", 0) > 3:
                return d
    raise RuntimeError(f"ドロワータイトルが '{section_heading}' に一致しない")


def _fb_close_drawer(page: Page) -> None:
    drawer = page.locator(FB_DRAWER_SEL).first
    if drawer.count() == 0:
        return
    cancel = page.locator(f"{FB_DRAWER_SEL} button").filter(has_text="キャンセル").first
    if cancel.count():
        cancel.click(force=True)
        page.wait_for_timeout(1200)


def _fb_save_drawer(page: Page, drawer: Locator) -> None:
    save_btn = drawer.locator("button").filter(has_text="保存する").first
    if save_btn.count() == 0:
        raise RuntimeError("保存するボタン未検出")
    save_btn.scroll_into_view_if_needed()
    with page.expect_response(
        lambda r: (
            FB_UPDATE_API_RE.search(r.url) is not None
            and r.request.method in ("POST", "PUT", "PATCH")
        ),
        timeout=15000,
    ) as resp_info:
        save_btn.click(force=True)
    resp = resp_info.value
    if resp.status >= 400:
        raise RuntimeError(f"update API {resp.status}: {resp.url}")
    page.wait_for_timeout(1000)


def _fb_set_radio(drawer: Locator, name: str, value: str) -> None:
    sel = f"input[type='radio'][name='{name}'][value='{value}']"
    el = drawer.locator(sel).first
    if el.count() == 0:
        raise RuntimeError(f"radio {name}={value} 未検出")
    el.evaluate(
        """el => {
            el.checked = true;
            el.dispatchEvent(new Event('input', {bubbles:true}));
            el.dispatchEvent(new Event('change', {bubbles:true}));
            el.click();
        }"""
    )


def _fb_set_select_by_options(drawer: Locator, option_texts: list[str], value: str) -> None:
    """ドロワー内のすべての SELECT を走査し、option.text に指定語のいずれかを含む SELECT を
    特定して、指定 value を選択する. 複数ステータスSELECTを区別するための工夫."""
    selects = drawer.locator("select")
    for i in range(selects.count()):
        s = selects.nth(i)
        try:
            option_list = s.evaluate(
                "el => Array.from(el.options).map(o => o.text)"
            )
        except Exception:
            continue
        if any(any(t in opt for opt in option_list) for t in option_texts):
            s.select_option(value=value)
            return
    raise RuntimeError(f"SELECT 未特定 (options hint={option_texts}, value={value})")


def fill_fb_management(fb_page: Page, internal_id: int) -> None:
    """管理情報ドロワーを開き、必要フィールドを設定して保存する."""
    drawer = _fb_open_section(fb_page, "管理情報")

    # ラジオ系
    _fb_set_radio(drawer, "potential_flg", "0")  # 掘り起こし: 対象外
    _fb_set_radio(drawer, "candidate_rank_id", "5")  # 人材ランク: E
    _fb_set_radio(drawer, "Sales_Type", "New")  # 営業種別: 新規
    _fb_set_radio(drawer, "traffic_source", "techdirect")  # 流入経路: TechDirect

    # SELECT 系: 集客ステータス / 営業ステータス / 自社担当者1
    # 識別はオプション文字列の特徴語で行う
    _fb_set_select_by_options(
        drawer, ["自動連絡", "面談設定済"], FB_RECRUIT_STATUS_VALUE
    )  # 集客ステータス (個別連絡=3)
    _fb_set_select_by_options(
        drawer, ["営業終了", "取引停止"], FB_SALES_STATUS_VALUE
    )  # 営業ステータス (営業不可=3)
    _fb_set_select_by_options(
        drawer, ["Claude アカウント"], FB_CLAUDE_ACCOUNT_ID
    )  # 自社担当者1 (Claude アカウント=840)

    _fb_save_drawer(fb_page, drawer)
    _fb_close_drawer(fb_page)


def fill_fb_comment(fb_page: Page, internal_id: int, comment_text: str) -> None:
    """候補者詳細の「コメント」タブ → 「コメントを作成」モーダルでコメント投稿する."""
    # コメントタブに遷移
    tab = fb_page.locator("a[role='tab'][href='#comment-tab']").first
    if tab.count() == 0:
        tab = fb_page.locator("a:has-text('コメント')").first
    if tab.count() == 0:
        raise RuntimeError("コメントタブ未検出")
    tab.scroll_into_view_if_needed()
    tab.click()
    fb_page.wait_for_timeout(1500)

    # 「コメントを作成」ボタン
    create_btn = fb_page.locator("button:has-text('コメントを作成')").first
    if create_btn.count() == 0:
        raise RuntimeError("『コメントを作成』ボタン未検出")
    create_btn.click()
    fb_page.wait_for_timeout(1500)

    # コメント入力モーダル内の textarea
    ta = fb_page.locator("textarea[placeholder='ここにコメントを記入してください']").first
    if ta.count() == 0:
        raise RuntimeError("コメント入力 textarea 未検出")
    ta.evaluate(
        """(el, v) => {
            el.focus();
            el.value = v;
            el.dispatchEvent(new Event('input', {bubbles:true}));
            el.dispatchEvent(new Event('change', {bubbles:true}));
            el.dispatchEvent(new Event('blur', {bubbles:true}));
        }""",
        comment_text,
    )
    fb_page.wait_for_timeout(500)

    # 投稿ボタン
    post_btn = fb_page.locator("button:has-text('コメントを投稿')").first
    if post_btn.count() == 0:
        raise RuntimeError("『コメントを投稿』ボタン未検出")
    # disabled が解除されるまで待つ
    for _ in range(20):
        if post_btn.is_enabled():
            break
        fb_page.wait_for_timeout(200)

    # API 応答を待つ（URL パスのヒントがないので POST 全般 + 閉じ判定の二段構え）
    posted_status: int | None = None

    def _capture(resp):
        nonlocal posted_status
        try:
            if (
                "/api/enterprise/" in resp.url
                and resp.request.method in ("POST", "PUT")
            ):
                posted_status = resp.status
                log(f"[INFO] comment related API {resp.status} ({resp.request.method} {resp.url})")
        except Exception:
            pass

    fb_page.on("response", _capture)
    try:
        post_btn.click(force=True)
        # モーダル が閉じる（コメント欄テキストが消える）まで待つ
        for _ in range(40):
            fb_page.wait_for_timeout(250)
            try:
                if fb_page.locator(
                    "textarea[placeholder='ここにコメントを記入してください']"
                ).count() == 0:
                    break
            except Exception:
                break
    finally:
        fb_page.remove_listener("response", _capture)

    if posted_status is not None and posted_status >= 400:
        _save_debug_capture(fb_page, "comment_fail")
        raise RuntimeError(f"comment API status={posted_status}")
    fb_page.wait_for_timeout(1500)


# ========== TechDirect ラベル付与 ==========

def apply_td_label(td_page: Page, detail_url: str) -> None:
    """TechDirect 候補者詳細で「リスト」ドロップダウンを開き、
    Base登録済みラベルのチェックを有効化する."""
    if td_page.url.rstrip("/") != detail_url.rstrip("/"):
        td_page.goto(detail_url, wait_until="networkidle")
        td_page.wait_for_timeout(2000)

    btn = td_page.locator("button:has-text('リスト')").first
    if btn.count() == 0:
        raise RuntimeError("『リスト』ボタン未検出")
    btn.click()
    td_page.wait_for_timeout(800)

    # JS で label.custom-control-label を辿り input をクリックして有効化
    result = td_page.evaluate(
        """(targetText) => {
            const labels = Array.from(document.querySelectorAll('label.custom-control-label'));
            for (const lb of labels) {
                const text = (lb.innerText || '').trim();
                if (text === targetText) {
                    const id = lb.getAttribute('for');
                    if (!id) continue;
                    const input = document.getElementById(id);
                    if (!input) continue;
                    if (!input.checked) {
                        input.click();
                    }
                    return { found: true, nowChecked: !!input.checked };
                }
            }
            return { found: false };
        }""",
        REGISTERED_LABEL,
    )
    td_page.wait_for_timeout(800)
    td_page.keyboard.press("Escape")
    td_page.wait_for_timeout(500)

    if not result.get("found"):
        raise RuntimeError(f"ラベル '{REGISTERED_LABEL}' がドロップダウンに見つからない")
    if not result.get("nowChecked"):
        raise RuntimeError(f"ラベル '{REGISTERED_LABEL}' のチェック有効化に失敗")


# ========== メインフロー ==========

def _save_refs_json(refs: list[CandidateRef]) -> Path:
    path = OUT_DIR / f"refs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    path.write_text(
        json.dumps([asdict(r) for r in refs], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    log(f"[INFO] 収集結果保存: {path} ({len(refs)} 件)")
    return path


def _load_refs_json(path: Path) -> list[CandidateRef]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return [CandidateRef(**d) for d in data]


def run(
    dry_run: bool,
    limit: int | None = None,
    only_url: str | None = None,
    fb_id_override: int | None = None,
    parallel: int = 1,
    refs_from: str | None = None,
    continue_on_error: bool = False,
    register_parallel: int = 1,
    full_refresh: bool = False,
) -> int:
    mode = "DRY-RUN" if dry_run else "本番"
    log(
        f"[START] techdirect-register-unsellable ({mode}) limit={limit} only_url={only_url} "
        f"fb_id_override={fb_id_override} parallel={parallel} refs_from={refs_from} "
        f"continue_on_error={continue_on_error} register_parallel={register_parallel} "
        f"full_refresh={full_refresh}"
    )

    # 処理済みURLキャッシュをロード。
    # --full-refresh は「ルックアップをスキップする」のみで、既存エントリは維持する
    # （refs_from などで一部URLのみ処理するケースでも、範囲外のキャッシュを失わないため）
    processed_cache: dict[str, dict] = load_processed_cache()
    log(f"[INFO] 処理済みURLキャッシュ: {len(processed_cache)} 件 ({PROCESSED_CACHE_PATH.name})")
    if full_refresh:
        log("[INFO] --full-refresh: キャッシュルックアップを無効化、既存エントリは維持")
    processed_cache_lock = threading.Lock()

    with sync_playwright() as pw:
        # TechDirect login（セッションキャッシュ優先）
        td_browser, td_context, td_page = td_login_with_cache(pw, headless=True)
        td_context.set_default_timeout(45000)

        # FreelanceBase login は本番モード時のみ（dry-runなら不要）
        fb_browser = fb_page = None
        if not dry_run:
            fb_browser, fb_context, fb_page = fb_login_with_cache(pw, headless=True)
            fb_context.set_default_timeout(45000)

        # --only-url 指定時は収集をスキップし、そのURLだけを処理
        if only_url:
            all_refs = [
                CandidateRef(
                    source_key="manual",
                    source_label="manual",
                    detail_url=only_url,
                    display_name="",
                )
            ]
            log(f"[INFO] --only-url 指定: {only_url}")
        elif refs_from:
            # --refs-from で事前保存した候補者リストを読み込み、収集フェーズをスキップ
            all_refs = _load_refs_json(Path(refs_from))
            log(f"[INFO] --refs-from 指定: {refs_from} から {len(all_refs)} 件読み込み")
        else:
            # --limit 時は収集も打ち切る（全ソース合計がlimitを超えたら追加収集しない）
            collect_cap_per_source = None
            if limit is not None and limit > 0:
                # limit の 3 倍を上限として各ソース単位に適用
                collect_cap_per_source = max(limit * 3, 20)

            all_refs = []
            for src in SOURCES:
                try:
                    if src["mode"] == "infinite_scroll":
                        refs = collect_from_applications(td_page, src["url"])
                    else:
                        refs = collect_from_scout_list(
                            td_page, src["url"], src["key"], src["label"],
                            collect_cap=collect_cap_per_source,
                        )
                    all_refs.extend(refs)
                except Exception as exc:
                    tb = traceback.format_exc()
                    log(f"[ERR] 収集失敗 {src['label']}: {exc}\n{tb}")
                    notify_slack(
                        f":rotating_light: register_unsellable ({mode}) 収集エラー\n"
                        f"ソース: {src['label']}\n```\n{tb[-1500:]}\n```"
                    )
                    td_browser.close()
                    if fb_browser:
                        fb_browser.close()
                    return 2

        log(f"[INFO] 全ソース合計（重複排除前）: {len(all_refs)} 件")

        # source_key 横断で重複排除（同じ候補者が複数リストにまたがる可能性）
        dedup: dict[str, CandidateRef] = {}
        for r in all_refs:
            m = re.search(r"/users/([0-9a-f-]{36})", r.detail_url)
            uid = m.group(1) if m else r.detail_url
            if uid not in dedup:
                dedup[uid] = r
        refs = list(dedup.values())
        log(f"[INFO] 重複排除後: {len(refs)} 件")

        if limit is not None and limit > 0:
            refs = refs[:limit]
            log(f"[INFO] --limit 適用後: {len(refs)} 件")

        # --only-url / --refs-from 以外のケースでは、収集結果を JSON 保存してリトライ可能にする
        if not only_url and not refs_from:
            _save_refs_json(refs)

        # キャッシュ適用: 終状態エントリはキャッシュから Judgment を復元し、
        # 判定フェーズからは除外する（詳細ページアクセスを省略）。
        # --full-refresh の場合はこのルックアップ自体をスキップ（全件再判定）。
        cached_judgments: list[Judgment] = []
        refs_to_judge: list[CandidateRef] = []
        for r in refs:
            entry = processed_cache.get(r.detail_url)
            if not full_refresh and entry and is_terminal_cache_entry(entry):
                cached_judgments.append(
                    Judgment(
                        ref=r,
                        status=entry.get("status", "skip_private"),
                        last_name=entry.get("last_name"),
                        first_name=entry.get("first_name"),
                        reason=f"cached ({entry.get('last_judged_at', '')})",
                    )
                )
            else:
                refs_to_judge.append(r)
        if cached_judgments:
            log(
                f"[INFO] キャッシュヒット: {len(cached_judgments)} 件スキップ "
                f"/ 新規判定対象: {len(refs_to_judge)} 件"
            )

        # 判定
        judgments_new: list[Judgment] = []
        if parallel <= 1:
            for i, ref in enumerate(refs_to_judge, start=1):
                log(f"[JUDGE {i}/{len(refs_to_judge)}] {ref.detail_url}")
                j = judge_candidate(td_page, ref)
                judgments_new.append(j)
                log(f"  → {j.status} ({j.reason})")
        elif not refs_to_judge:
            log("[INFO] 判定対象0件（全てキャッシュヒット）")
        else:
            total = len(refs_to_judge)
            log(f"[INFO] 並列判定 N={parallel} / 対象 {total} 件")
            # 進捗カウンタ（スレッドセーフ）
            counter_lock = threading.Lock()
            counter = {"n": 0}
            # 順序を保つため index 付きで判定
            indexed_refs = list(enumerate(refs_to_judge))
            results_by_index: dict[int, Judgment] = {}

            def _judge_thread(pair: tuple[int, CandidateRef]) -> tuple[int, Judgment]:
                idx, ref = pair
                page = _get_worker_page()
                j = judge_candidate(page, ref)
                with counter_lock:
                    counter["n"] += 1
                    n = counter["n"]
                log(f"[JUDGE {n}/{total}] {ref.detail_url} → {j.status} ({j.reason})")
                return idx, j

            try:
                with ThreadPoolExecutor(max_workers=parallel) as ex:
                    for idx, j in ex.map(_judge_thread, indexed_refs):
                        results_by_index[idx] = j
            finally:
                _cleanup_workers()

            # 入力順で並べる（結果の順序をシリアル版と一致させる）
            judgments_new = [results_by_index[i] for i in range(total)]

        # キャッシュヒット分と新規判定分を結合（入力順を維持）
        judge_by_url = {j.ref.detail_url: j for j in cached_judgments + judgments_new}
        judgments: list[Judgment] = [judge_by_url.get(r.detail_url) for r in refs if judge_by_url.get(r.detail_url)]

        # 新規判定結果をキャッシュに反映（登録フェーズ前に save）
        now = datetime.now().isoformat(timespec="seconds")
        for j in judgments_new:
            entry = processed_cache.get(j.ref.detail_url, {})
            entry.update({
                "status": j.status,
                "last_name": j.last_name,
                "first_name": j.first_name,
                "last_judged_at": now,
            })
            if j.error:
                entry["last_judge_error"] = j.error[:300]
            processed_cache[j.ref.detail_url] = entry
        if judgments_new:
            save_processed_cache(processed_cache)
            log(f"[INFO] 判定フェーズ完了、キャッシュ保存: {PROCESSED_CACHE_PATH.name}")

        # 本番モード: valid 判定された候補者について FB 登録＋TD ラベル付与
        registrations: list[RegistrationResult] = []
        if not dry_run:
            valid_js = [j for j in judgments if j.status == "valid"]
            log(f"[INFO] 登録対象: {len(valid_js)} 件")

            # 並列登録ブランチ（register_parallel > 1 の場合）
            if register_parallel > 1 and fb_id_override is None:
                # バッチ内 pre-dedup: 同姓同名は先頭のみ残し、以降は batch_dup として即スキップ
                # （並列重複チェックのレース条件で同じ名前の重複が作られる事故を防ぐ）
                seen: dict[tuple, Judgment] = {}
                batch_dups: list[Judgment] = []
                for j in valid_js:
                    key = (j.last_name or "", j.first_name or "")
                    if key in seen:
                        batch_dups.append(j)
                    else:
                        seen[key] = j
                primary_valids = list(seen.values())
                if batch_dups:
                    log(
                        f"[INFO] バッチ内 pre-dedup: primary={len(primary_valids)} / "
                        f"batch_dup={len(batch_dups)}"
                    )
                    for j in batch_dups:
                        result = RegistrationResult(judgment=j)
                        result.skipped_duplicate = True
                        result.error = (
                            f"batch_dup: 同バッチ内で先行登録済み ({j.last_name} {j.first_name})"
                        )
                        registrations.append(result)

                total_reg = len(primary_valids)
                valid_js_for_parallel = primary_valids
                log(f"[INFO] 並列登録 N={register_parallel} / 対象 {total_reg} 件")
                reg_counter_lock = threading.Lock()
                reg_counter = {"n": 0}

                def _register_thread(pair: tuple[int, Judgment]) -> RegistrationResult:
                    idx, j = pair
                    fb_page_w, td_page_w = _get_worker_reg_pages()
                    result = RegistrationResult(judgment=j)
                    try:
                        existing = check_fb_duplicate(
                            fb_page_w, j.last_name or "", j.first_name or ""
                        )
                        if existing:
                            result.skipped_duplicate = True
                            result.duplicate_ids = existing
                            # 既存FBに同姓同名があり登録スキップは「終状態」なので
                            # キャッシュに記録し、次回以降は詳細アクセスすらスキップする
                            with processed_cache_lock:
                                entry = processed_cache.get(j.ref.detail_url, {})
                                entry.update({
                                    "status": "skip_duplicate",
                                    "last_name": j.last_name,
                                    "first_name": j.first_name,
                                    "duplicate_ids": existing,
                                    "last_judged_at": datetime.now().isoformat(timespec="seconds"),
                                })
                                processed_cache[j.ref.detail_url] = entry
                            with reg_counter_lock:
                                reg_counter["n"] += 1
                                done = reg_counter["n"]
                            log(
                                f"[REGISTER {idx}/{total_reg}] ({done} done) "
                                f"dup={existing}: {j.ref.detail_url}"
                            )
                            return result
                        internal_id, email = create_fb_candidate_with_retry(
                            fb_page_w, j.last_name or "", j.first_name or ""
                        )
                        result.fb_email = email
                        result.fb_internal_id = internal_id
                        fill_fb_management(fb_page_w, internal_id)
                        fill_fb_comment(
                            fb_page_w,
                            internal_id,
                            COMMENT_TEMPLATE.format(detail_url=j.ref.detail_url),
                        )
                        result.registered = True
                        apply_td_label(td_page_w, j.ref.detail_url)
                        result.label_applied = True
                        # キャッシュ更新（並列書き込みは lock で保護）
                        with processed_cache_lock:
                            entry = processed_cache.get(j.ref.detail_url, {})
                            entry.update({
                                "status": "valid",
                                "last_name": j.last_name,
                                "first_name": j.first_name,
                                "fb_internal_id": internal_id,
                                "label_applied": True,
                                "last_registered_at": datetime.now().isoformat(timespec="seconds"),
                            })
                            processed_cache[j.ref.detail_url] = entry
                        with reg_counter_lock:
                            reg_counter["n"] += 1
                            done = reg_counter["n"]
                        log(
                            f"[OK {idx}/{total_reg}] ({done} done) "
                            f"internal_id={internal_id} {j.last_name} {j.first_name}"
                        )
                    except Exception as exc:
                        tb = traceback.format_exc()
                        result.error = str(exc)
                        with reg_counter_lock:
                            reg_counter["n"] += 1
                            done = reg_counter["n"]
                        log(
                            f"[ERR {idx}/{total_reg}] ({done} done) "
                            f"{j.ref.detail_url} → {j.last_name} {j.first_name}: {exc}"
                        )
                        # ドロワー等の残骸を閉じる
                        try:
                            _fb_close_drawer(fb_page_w)
                        except Exception:
                            pass
                        try:
                            fb_page_w.keyboard.press("Escape")
                        except Exception:
                            pass
                        notify_slack(
                            f":warning: register_unsellable 登録失敗（並列・継続中）\n"
                            f"候補者: {j.ref.detail_url}\n"
                            f"氏名: {j.last_name} {j.first_name}\n"
                            f"FB internal_id: {result.fb_internal_id}\n"
                            f"登録済: {result.registered} / "
                            f"ラベル付与: {result.label_applied}\n"
                            f"```\n{tb[-800:]}\n```"
                        )
                    return result

                indexed_valids = list(enumerate(valid_js_for_parallel, start=1))
                try:
                    with ThreadPoolExecutor(max_workers=register_parallel) as ex:
                        for r in ex.map(_register_thread, indexed_valids):
                            registrations.append(r)
                finally:
                    _cleanup_reg_workers()

                # 並列登録ブランチは下の逐次ループをスキップ
                valid_js = []

            for i, j in enumerate(valid_js, start=1):
                log(f"[REGISTER {i}/{len(valid_js)}] {j.ref.detail_url} → {j.last_name} {j.first_name}")
                result = RegistrationResult(judgment=j)
                try:
                    if fb_id_override is not None:
                        log(f"[INFO] --fb-id 指定: {fb_id_override} を使い、新規作成と重複チェックをスキップ")
                        internal_id = fb_id_override
                        result.fb_email = None
                        # 既存レコードの詳細ページへ明示遷移
                        fb_page.goto(
                            f"https://freelancebase.jp/enterprise/candidates/{internal_id}",
                            wait_until="networkidle",
                        )
                        fb_page.wait_for_timeout(2500)
                    else:
                        # 同姓同名チェック（FB）。既存があればスキップ
                        existing = check_fb_duplicate(
                            fb_page, j.last_name or "", j.first_name or ""
                        )
                        if existing:
                            log(
                                f"[INFO] FBに同姓同名の既存候補者あり: {existing} → "
                                "登録・ラベル付与をスキップ"
                            )
                            result.skipped_duplicate = True
                            result.duplicate_ids = existing
                            # 終状態扱いでキャッシュ更新（次回以降は詳細アクセス省略）
                            entry = processed_cache.get(j.ref.detail_url, {})
                            entry.update({
                                "status": "skip_duplicate",
                                "last_name": j.last_name,
                                "first_name": j.first_name,
                                "duplicate_ids": existing,
                                "last_judged_at": datetime.now().isoformat(timespec="seconds"),
                            })
                            processed_cache[j.ref.detail_url] = entry
                            registrations.append(result)
                            continue
                        internal_id, email = create_fb_candidate_with_retry(
                            fb_page, j.last_name or "", j.first_name or ""
                        )
                        result.fb_email = email
                    result.fb_internal_id = internal_id

                    # 管理情報入力
                    fill_fb_management(fb_page, internal_id)

                    # コメントタブに TechDirect URL とメモを投稿
                    comment_body = COMMENT_TEMPLATE.format(detail_url=j.ref.detail_url)
                    fill_fb_comment(fb_page, internal_id, comment_body)

                    result.registered = True

                    # TechDirect 側でラベル付与
                    apply_td_label(td_page, j.ref.detail_url)
                    result.label_applied = True

                    # キャッシュ更新（逐次）
                    entry = processed_cache.get(j.ref.detail_url, {})
                    entry.update({
                        "status": "valid",
                        "last_name": j.last_name,
                        "first_name": j.first_name,
                        "fb_internal_id": internal_id,
                        "label_applied": True,
                        "last_registered_at": datetime.now().isoformat(timespec="seconds"),
                    })
                    processed_cache[j.ref.detail_url] = entry

                    log(f"[OK] 登録完了 (internal_id={internal_id})")
                except Exception as exc:
                    tb = traceback.format_exc()
                    result.error = str(exc)
                    log(f"[ERR] 登録失敗: {exc}\n{tb}")
                    registrations.append(result)
                    if continue_on_error:
                        # 失敗は記録するが続行。UI/ページ状態が不整合の可能性があるため
                        # ドロワー系を閉じてから次へ
                        try:
                            _fb_close_drawer(fb_page)
                        except Exception:
                            pass
                        try:
                            fb_page.keyboard.press("Escape")
                        except Exception:
                            pass
                        notify_slack(
                            f":warning: register_unsellable 登録失敗（継続中）\n"
                            f"候補者: {j.ref.detail_url}\n"
                            f"氏名: {j.last_name} {j.first_name}\n"
                            f"FB internal_id: {result.fb_internal_id}\n"
                            f"登録済: {result.registered} / ラベル付与: {result.label_applied}\n"
                            f"```\n{tb[-800:]}\n```"
                        )
                        continue
                    notify_slack(
                        f":rotating_light: register_unsellable 登録失敗（即中断）\n"
                        f"候補者: {j.ref.detail_url}\n"
                        f"氏名: {j.last_name} {j.first_name}\n"
                        f"FB internal_id: {result.fb_internal_id}\n"
                        f"登録済: {result.registered} / ラベル付与: {result.label_applied}\n"
                        f"```\n{tb[-1500:]}\n```"
                    )
                    td_browser.close()
                    if fb_browser:
                        fb_browser.close()
                    # 結果とキャッシュを保存してから終了
                    _save_result(mode, judgments, registrations)
                    save_processed_cache(processed_cache)
                    return 3
                registrations.append(result)

        td_browser.close()
        if fb_browser:
            fb_browser.close()

    # 集計
    summary: dict[str, int] = {}
    for j in judgments:
        summary[j.status] = summary.get(j.status, 0) + 1

    # 登録フェーズで更新したキャッシュを最終保存
    save_processed_cache(processed_cache)
    log(f"[INFO] 最終キャッシュ保存: {PROCESSED_CACHE_PATH.name} ({len(processed_cache)} 件)")

    result_path = _save_result(mode, judgments, registrations)

    # Slack通知
    valid = [j for j in judgments if j.status == "valid"]
    summary_lines = [f"- {k}: {v} 件" for k, v in sorted(summary.items())]
    emoji = ":white_check_mark:"
    reg_line = ""
    if not dry_run:
        registered_count = sum(1 for r in registrations if r.registered and r.label_applied)
        failed_count = sum(1 for r in registrations if r.error)
        reg_line = (
            f"\n本番登録成功: {registered_count} 件 / 失敗: {failed_count} 件"
        )
        if failed_count > 0:
            emoji = ":warning:"

    # 重複スキップの詳細を Slack 本文に含める
    dup_lines = []
    if not dry_run:
        dup_results = [r for r in registrations if r.skipped_duplicate]
        if dup_results:
            emoji = ":warning:"
            dup_lines.append(f"\n⚠️ FB同姓同名で登録スキップ: {len(dup_results)} 件")
            for r in dup_results[:10]:  # 最大10件
                dup_lines.append(
                    f"  • {r.judgment.last_name} {r.judgment.first_name} — "
                    f"既存FB ID={r.duplicate_ids} / TD={r.judgment.ref.detail_url}"
                )

    slack_text = (
        f"{emoji} register_unsellable {mode} 完了\n"
        f"対象合計: {len(judgments)} 件\n"
        + "\n".join(summary_lines) + "\n"
        + f"登録対象（valid）: {len(valid)} 件"
        + reg_line
        + "".join(dup_lines) + "\n"
        + f"結果ファイル: `{result_path.name}`"
    )
    notify_slack(slack_text)

    log("[DONE]")
    return 0


def _save_result(
    mode: str,
    judgments: list[Judgment],
    registrations: list[RegistrationResult],
) -> Path:
    prefix = "dryrun" if mode == "DRY-RUN" else "run"
    result_path = OUT_DIR / f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    summary: dict[str, int] = {}
    for j in judgments:
        summary[j.status] = summary.get(j.status, 0) + 1
    result_path.write_text(
        json.dumps(
            {
                "mode": mode,
                "timestamp": datetime.now().isoformat(),
                "summary": summary,
                "judgments": [
                    {**asdict(j), "ref": asdict(j.ref)} for j in judgments
                ],
                "registrations": [
                    {
                        "detail_url": r.judgment.ref.detail_url,
                        "last_name": r.judgment.last_name,
                        "first_name": r.judgment.first_name,
                        "registered": r.registered,
                        "fb_internal_id": r.fb_internal_id,
                        "fb_email": r.fb_email,
                        "label_applied": r.label_applied,
                        "error": r.error,
                        "skipped_duplicate": r.skipped_duplicate,
                        "duplicate_ids": r.duplicate_ids,
                    }
                    for r in registrations
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    log(f"[INFO] 結果保存: {result_path}")
    return result_path


def main() -> int:
    ap = argparse.ArgumentParser(description="TechDirect営業対象外人材をBase登録")
    ap.add_argument("--dry-run", action="store_true", help="収集＋氏名判定のみ（本番書込なし）")
    ap.add_argument("--limit", type=int, default=None, help="判定対象を先頭N件に制限（スモークテスト用）")
    ap.add_argument(
        "--only-url",
        default=None,
        help="指定したTechDirect人材詳細URLのみを判定・登録する（収集フェーズをスキップ）",
    )
    ap.add_argument(
        "--fb-id",
        type=int,
        default=None,
        help="FreelanceBase の新規作成をスキップし、既存の internal_id に対して管理情報入力とラベル付与のみ実行（デバッグ用）",
    )
    ap.add_argument(
        "--parallel",
        type=int,
        default=1,
        help="判定フェーズの並列数（1=従来通り、推奨3〜5）",
    )
    ap.add_argument(
        "--refs-from",
        default=None,
        help="収集済みの候補者リストJSONから判定を開始（収集フェーズをスキップ）",
    )
    ap.add_argument(
        "--continue-on-error",
        action="store_true",
        help="登録フェーズで個別候補者の失敗が発生しても即中断せず、残りの処理を続行する",
    )
    ap.add_argument(
        "--register-parallel",
        type=int,
        default=1,
        help="登録フェーズの並列数（1=従来通り、推奨2〜3）. 並列時は continue_on_error 相当の挙動",
    )
    ap.add_argument(
        "--full-refresh",
        action="store_true",
        help="processed_urls.json のキャッシュを無視して全件再判定。四半期に1度 or 問題発生時に実行",
    )
    args = ap.parse_args()

    try:
        return run(
            dry_run=args.dry_run,
            limit=args.limit,
            only_url=args.only_url,
            fb_id_override=args.fb_id,
            parallel=args.parallel,
            refs_from=args.refs_from,
            continue_on_error=args.continue_on_error,
            register_parallel=args.register_parallel,
            full_refresh=args.full_refresh,
        )
    except Exception as exc:
        tb = traceback.format_exc()
        log(f"[FATAL] {exc}\n{tb}")
        notify_slack(
            f":rotating_light: register_unsellable 致命的エラー\n```\n{tb[-1800:]}\n```"
        )
        return 99


if __name__ == "__main__":
    sys.exit(main())
