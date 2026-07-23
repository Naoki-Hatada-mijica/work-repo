#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excelスキルシートの見た目（罫線・列幅・結合セル・配置・塗り・縦書き）を
保ったまま、読みやすいA4縦PDF用のHTMLに変換する汎用スクリプト。
中身は変えない。フォーマットが違うスキルシートでも動くよう自動判定する。

使い方: python3 xlsx_to_html.py <入力.xlsx> [シート名] [出力.html]
"""
import sys, html
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.datetime import from_excel

SRC = sys.argv[1]
SHEET = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2] else None
OUT = sys.argv[3] if len(sys.argv) > 3 else 'grid.html'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb[SHEET] if (SHEET and SHEET in wb.sheetnames) else wb.worksheets[0]

FONT = ('"Meiryo","メイリオ","游ゴシック","Yu Gothic",'
        '"Hiragino Kaku Gothic ProN","Hiragino Sans",sans-serif')
KEEP_TOGETHER = True   # 案件をページ途中で分割しない（ページに収まらない案件は次ページへ）
ROW_SCALE = 1.22       # 行の高さ（文字サイズは変えない）。1.0＝Excelどおり、大きいほど行間にゆとり
LINE_HEIGHT = 1.6      # 折り返し本文の行送り。ROW_SCALE とセットで調整する
CELL_PAD_Y = 2         # セル上下の余白(px)

# ===== 罫線の強弱 =====
# 外枠と案件の切り替わりは実線（太め）、案件の中は薄い点線にして、紙面を軽くする。
LIGHT_INNER_BORDER = True     # 案件の中の細い罫線を点線に落とす
INNER_BORDER = '1px dotted #000'      # 案件の中（黒の点線）
STRONG_MIN_PX = 1.2           # 外枠・案件の区切りは最低この太さの実線にする

# ===== 1. 印刷範囲を自動検出（印刷範囲指定→なければ内容＋罫線＋塗りの範囲）=====
def content_bounds():
    maxr = maxc = 1
    for row in ws.iter_rows():
        for c in row:
            has = c.value not in (None, '')
            if not has:
                b = c.border
                has = any(s and s.style for s in (b.top, b.bottom, b.left, b.right))
            if not has:
                fl = c.fill
                has = getattr(fl, 'patternType', None) == 'solid'
            if has:
                if c.row > maxr: maxr = c.row
                if c.column > maxc: maxc = c.column
    for rng in ws.merged_cells.ranges:
        maxr = max(maxr, rng.max_row); maxc = max(maxc, rng.max_col)
    return maxr, maxc

pa = ws.print_area
if pa:
    pa = pa[0] if isinstance(pa, (list, tuple)) else pa
    pa = pa.split('!')[-1].replace('$', '')
    MINC, MINR, MAXC, MAXR = range_boundaries(pa)  # 印刷範囲がA1始まりでない場合に対応
else:
    MINR, MINC = 1, 1
    MAXR, MAXC = content_bounds()

# 隠れ行・隠れ列（Excelで非表示のもの）は高さ／幅0にして畳む
HIDDEN_ROWS = {r for r, d in ws.row_dimensions.items() if d.hidden}
HIDDEN_COLS = set()
for c, d in ws.column_dimensions.items():
    if d.hidden:
        for ci in range(d.min, (d.max or d.min) + 1):
            HIDDEN_COLS.add(ci)

CORRECTIONS = {}  # 誤字補正はファイル固有なので既定では行わない

# ===== 本文フォントサイズの統一 =====
# 折り返し（本文）セルの中で最も多いサイズを「標準」とし、そこから極端に離れて
# いない本文セルは標準に合わせる（例：一部の案件だけ11ptなのを10ptに統一）。
# タイトルや大見出しなど大きく離れたサイズはそのまま維持する。
import collections as _collections
_body = _collections.Counter()
for _row in ws.iter_rows(min_row=MINR, max_row=MAXR, min_col=MINC, max_col=MAXC):
    for _c in _row:
        if (_c.alignment and _c.alignment.wrap_text and isinstance(_c.value, str)
                and len(_c.value) > 25 and _c.font and _c.font.sz):
            _body[_c.font.sz] += 1
BODY_SIZE = _body.most_common(1)[0][0] if _body else None

def norm_size(cell):
    """本文セルのフォントサイズを標準に寄せる。それ以外は元のサイズ。"""
    sz = cell.font.sz if (cell.font and cell.font.sz) else None
    if sz is None or BODY_SIZE is None:
        return sz
    al = cell.alignment
    if not (al and al.wrap_text):
        return sz
    v = cell.value
    # 長い本文（折り返しセル）は、サイズが大きく外れていても標準に統一する
    # （例：ある案件だけ業務内容が20pt→10ptへ）。短いセル（見出し等）は範囲内のみ。
    if isinstance(v, str) and len(v) > 30:
        return BODY_SIZE
    if 0.6 * BODY_SIZE <= sz <= 1.5 * BODY_SIZE:
        return BODY_SIZE
    return sz

# ===== 2. 案件（まとまり）の区切りを自動検出 =====
def is_idx(v):
    return isinstance(v, int) or (isinstance(v, str) and v.strip().isdigit())

def idx_rows(col):
    out = []
    for r in range(MINR, MAXR + 1):
        if r in HIDDEN_ROWS:
            continue
        v = ws.cell(r, col).value
        if is_idx(v) and 1 <= int(str(v).strip()) <= 99:
            out.append(r)
    return out

# スキル評価セクション（「■…スキル…」で始まる行）を先に検出
skill_start = None
for r in range(MINR, MAXR + 1):
    if any(isinstance(ws.cell(r, c).value, str) and '■' in ws.cell(r, c).value
           and 'スキル' in ws.cell(r, c).value for c in range(MINC, MINC + 3)):
        skill_start = r
        break

# (a) 案件番号ベース：先頭列／次列に 1〜99 の番号が並ぶ形式
a_rows = idx_rows(MINC)
num_starts = a_rows if len(a_rows) >= 2 else idx_rows(MINC + 1)

# (b) 見出しベース：業務内容タイトル（3列以上の結合セルで「■」始まり）。
#     番号のない案件（個人開発など）もこれで拾える。ヘッダやスキル欄は除外。
first_ref = min(num_starts) if num_starts else MINR
title_starts = []
for rng in ws.merged_cells.ranges:
    r0 = rng.min_row
    if (rng.max_col - rng.min_col + 1) >= 3 and MINR <= r0 <= MAXR and r0 >= first_ref \
            and not (skill_start and r0 >= skill_start):
        v = ws.cell(r0, rng.min_col).value
        if isinstance(v, str) and v.strip().startswith('■'):
            title_starts.append(r0)

# (a)(b) を統合し、ごく近接（2行以内＝番号行と見出し行の重複）だけ1つにまとめる
proj_starts = []
for r in sorted(set(num_starts) | set(title_starts)):
    if not proj_starts or r - proj_starts[-1] > 2:
        proj_starts.append(r)

# ブロック生成（先頭＝見出し等／各案件＝分割禁止／スキル評価＝分割可）
BLOCKS = []
if proj_starts:
    first = proj_starts[0]
    if first > MINR:
        BLOCKS.append((MINR, first - 1, False))
    end_proj = (skill_start - 1) if skill_start else MAXR
    for i, s in enumerate(proj_starts):
        e = proj_starts[i + 1] - 1 if i + 1 < len(proj_starts) else end_proj
        BLOCKS.append((s, e, True))
    if skill_start:
        BLOCKS.append((skill_start, MAXR, False))
else:
    BLOCKS = [(MINR, MAXR, False)]

# 罫線の強弱用：ブロックの境界行と、「案件の中」の行を控えておく。
# 点線にするのは案件ブロックの中だけ。見出し部（氏名・資格・スキル要約など）と
# スキル評価表は、箱ごとの区切りが見えたほうが読みやすいので Excel の罫線のまま。
BOUND_TOP = {r0 for (r0, _r1, _k) in BLOCKS} | {MINR}
BOUND_BOTTOM = {r1 for (_r0, r1, _k) in BLOCKS} | {MAXR}
INNER_ROWS = set()
for (r0, r1, keep) in BLOCKS:
    if keep:   # keep=True＝案件ブロック
        INNER_ROWS.update(range(r0, r1 + 1))

# 「業務内容」列の位置を探す。ここから右が案件の中身＝点線、
# 左にある項番(No)・期間の欄は実線（太線）にして、案件の枠組みをはっきりさせる。
CONTENT_COL = None
_hdr_end = proj_starts[0] if proj_starts else MAXR + 1
for _r in range(MINR, _hdr_end):
    for _c in range(MINC, MAXC + 1):
        _v = ws.cell(_r, _c).value
        if isinstance(_v, str) and any(k in _v for k in ('業務内容', '職務内容', '業務経歴')):
            CONTENT_COL = _c
            for _rng in ws.merged_cells.ranges:   # 結合セルなら左端の列を採用
                if _rng.min_row <= _r <= _rng.max_row and _rng.min_col <= _c <= _rng.max_col:
                    CONTENT_COL = _rng.min_col
                    break
            break
    if CONTENT_COL:
        break

# ===== 3. 列幅・行高 =====
DEFAULT_W = ws.sheet_format.defaultColWidth or 8.43
width_map = {}
for cd in ws.column_dimensions.values():
    if cd.width is None:
        continue
    for ci in range(cd.min, min(cd.max, MAXC) + 1):
        width_map[ci] = cd.width
def col_px(c):
    if c in HIDDEN_COLS:
        return 0
    return round(width_map.get(c, DEFAULT_W) * 7 + 5)
colw = [col_px(c) for c in range(MINC, MAXC + 1)]
table_w = sum(colw)

def row_px(r):
    if r in HIDDEN_ROWS:
        return 0
    rd = ws.row_dimensions.get(r)
    h = rd.height if (rd and rd.height) else 13.5
    return round(h * 96 / 72 * ROW_SCALE, 1)

# ===== 4. 結合セル =====
anchor, covered = {}, set()
for rng in ws.merged_cells.ranges:
    r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
    if r1 > MAXR or c1 > MAXC or r2 < MINR or c2 < MINC:
        continue
    r1, c1 = max(r1, MINR), max(c1, MINC)
    r2, c2 = min(r2, MAXR), min(c2, MAXC)
    anchor[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            if (rr, cc) != (r1, c1):
                covered.add((rr, cc))

# ===== 5. セルのスタイル =====
BSTYLE = {'thin': '1px solid', 'hair': '0.5px solid', 'medium': '1.5px solid',
          'thick': '2.5px solid', 'double': '3px double', 'dotted': '1px dotted',
          'dashed': '1px dashed', 'dashDot': '1px dashed', 'mediumDashed': '1.5px dashed',
          'slantDashDot': '1px dashed'}

# --- テーマ色パレット（theme1.xml の clrScheme を読む）---
# Excel の theme 番号 → 配色名の対応は 0/1 と 2/3 が入れ替わる（既知の仕様）
THEME_ORDER = ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3',
               'accent4', 'accent5', 'accent6', 'hlink', 'folHlink']
THEME_FALLBACK = {'lt1': 'FFFFFF', 'dk1': '000000', 'lt2': 'EEECE1', 'dk2': '1F497D',
                  'accent1': '4F81BD', 'accent2': 'C0504D', 'accent3': '9BBB59',
                  'accent4': '8064A2', 'accent5': '4BACC6', 'accent6': 'F79646',
                  'hlink': '0000FF', 'folHlink': '800080'}

def load_theme_colors():
    palette = dict(THEME_FALLBACK)
    raw = getattr(wb, 'loaded_theme', None)
    if not raw:
        return palette
    try:
        import xml.etree.ElementTree as ET
        ns = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
        root = ET.fromstring(raw)
        scheme = root.find(f'.//{ns}clrScheme')
        if scheme is None:
            return palette
        for node in scheme:
            name = node.tag.split('}')[-1]
            srgb = node.find(f'{ns}srgbClr')
            sysc = node.find(f'{ns}sysClr')
            if srgb is not None and srgb.get('val'):
                palette[name] = srgb.get('val')
            elif sysc is not None and sysc.get('lastClr'):
                palette[name] = sysc.get('lastClr')
    except Exception:
        pass
    return palette

THEME = load_theme_colors()

def apply_tint(hex6, tint):
    """Excel の濃淡（tint）を適用する。明度を白／黒方向へ寄せる仕様どおりの計算。"""
    if not tint:
        return hex6
    import colorsys
    r, g, b = (int(hex6[i:i + 2], 16) / 255 for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = l * (1 + tint) if tint < 0 else l * (1 - tint) + tint
    r, g, b = colorsys.hls_to_rgb(h, min(max(l, 0.0), 1.0), s)
    return '{:02X}{:02X}{:02X}'.format(*(round(v * 255) for v in (r, g, b)))

def argb(color):
    """セル色を #RRGGBB に解決する。RGB 直指定・テーマ色・インデックス色に対応。"""
    if color is None:
        return None
    ctype = getattr(color, 'type', None)
    tint = getattr(color, 'tint', 0.0) or 0.0
    base = None
    if ctype == 'theme':
        idx = getattr(color, 'theme', None)
        if isinstance(idx, int) and 0 <= idx < len(THEME_ORDER):
            base = THEME[THEME_ORDER[idx]]
    elif ctype == 'indexed':
        from openpyxl.styles.colors import COLOR_INDEX
        idx = getattr(color, 'indexed', None)
        if isinstance(idx, int) and 0 <= idx < len(COLOR_INDEX):
            v = COLOR_INDEX[idx]
            if isinstance(v, str) and len(v) == 8:
                base = v[2:]
    else:
        rgb = getattr(color, 'rgb', None)
        if isinstance(rgb, str) and len(rgb) == 8:
            base = rgb[2:]
    if not base:
        return None
    return '#' + apply_tint(base.upper(), tint)

THIN_STYLES = ('thin', 'hair', 'dotted', 'dashed', 'dashDot', 'slantDashDot')

def border_css(side, strong=True, inner=False):
    """罫線をCSSにする。
      strong=True  … 外枠・案件の区切り。細くても最低 STRONG_MIN_PX の実線にする
      inner=True   … 案件の中。細い罫線は薄い点線に落として紙面を軽くする
    """
    if not side or not side.style:
        return '0'
    if inner and LIGHT_INNER_BORDER and side.style in THIN_STYLES:
        return INNER_BORDER
    css = BSTYLE.get(side.style, '1px solid')
    color = argb(side.color) or '#000'
    if strong and side.style in ('thin', 'hair'):
        css = f'{STRONG_MIN_PX}px solid'
    return f'{css} {color}'

def cell_style(cell, r=None, c=None, span=(1, 1)):
    s = []
    b = cell.border
    if r is None:
        r, c = cell.row, cell.column
    rb = r + span[0] - 1          # このセルの下端の行
    cr = c + span[1] - 1          # このセルの右端の列
    # 点線にするのは案件ブロックの中で、かつ「業務内容」列から右だけ。
    # 項番(No)・期間の欄は案件の中でも実線（太線）のまま残す。
    in_proj = (r in INNER_ROWS
               and (CONTENT_COL is None or c >= CONTENT_COL))
    idx_area = (r in INNER_ROWS and CONTENT_COL is not None and cr < CONTENT_COL)
    # 外枠・案件の切り替わり・項番/期間欄＝実線（太め）、案件の中身＝点線
    top_strong = (r in BOUND_TOP) or idx_area
    bot_strong = (rb in BOUND_BOTTOM) or idx_area
    left_strong = (c == MINC) or idx_area
    right_strong = (cr == MAXC) or idx_area or (CONTENT_COL is not None and cr == CONTENT_COL - 1)
    s.append(f'border-top:{border_css(b.top, top_strong, in_proj and not top_strong)}')
    s.append(f'border-bottom:{border_css(b.bottom, bot_strong, in_proj and not bot_strong)}')
    s.append(f'border-left:{border_css(b.left, left_strong, in_proj and not left_strong)}')
    s.append(f'border-right:{border_css(b.right, right_strong, in_proj and not right_strong)}')
    f = cell.font
    if f:
        sz = norm_size(cell)  # 本文サイズを標準に統一
        if sz:
            s.append(f'font-size:{round(sz * 96 / 72)}px')  # pt→px
        if f.bold:
            s.append('font-weight:700')
        fc = argb(f.color)
        if fc and fc != '#000000':
            s.append(f'color:{fc}')
    fl = cell.fill
    if fl and getattr(fl, 'patternType', None) == 'solid':
        bg = argb(fl.fgColor)
        if bg:
            s.append(f'background:{bg}')
    al = cell.alignment
    s.append('vertical-align:middle')
    if al and al.textRotation == 255:  # 縦書き
        s.append('writing-mode:vertical-rl; text-orientation:upright; white-space:nowrap')
        s.append('text-align:center')
        return ';'.join(s)
    h = (al.horizontal if al else None) or \
        ('right' if isinstance(cell.value, (int, float)) else 'left')
    s.append(f'text-align:{h}')
    if al and al.wrap_text:
        s.append('white-space:pre-wrap; word-break:break-word')
    else:
        s.append('white-space:nowrap')
    return ';'.join(s)

def cell_text(r, c, cell):
    val = CORRECTIONS.get((r, c), cell.value)
    if val is None:
        return ''
    # 日付シリアル値の対応：
    #  ・日付書式が付いている数値、または
    #  ・2000〜2049年相当（36526〜54789）の裸の数値（＝書式漏れの日付とみなす）
    # のみ日付へ変換する。それ以外の数値はそのまま。
    nf = cell.number_format or ''
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        is_date_fmt = any(t in nf for t in ('y', 'm', 'd', '年', '月', '日'))
        if (is_date_fmt and val > 59) or (36526 <= val <= 54789):
            try:
                val = from_excel(val)
            except Exception:
                pass
    if hasattr(val, 'strftime'):
        return val.strftime('%Y年%-m月')
    return html.escape(str(val)).replace('\n', '<br>')

# ===== 6. HTML生成 =====
COLGROUP = '<colgroup>' + ''.join(f'<col style="width:{w}px">' for w in colw) + '</colgroup>'

def render_rows(r0, r1):
    out = []
    for r in range(r0, r1 + 1):
        out.append(f'<tr style="height:{row_px(r)}px">')
        for c in range(MINC, MAXC + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(r, c)
            span = anchor.get((r, c), (1, 1))
            attrs = ''
            if span[0] > 1:
                attrs += f' rowspan="{span[0]}"'
            if span[1] > 1:
                attrs += f' colspan="{span[1]}"'
            style = cell_style(cell, r, c, span)
            # セルが占める行がすべて非表示なら中身を畳む（高さを持たせない）
            if sum(row_px(rr) for rr in range(r, r + span[0])) == 0:
                style += ';font-size:0;line-height:0;padding:0'
                txt = ''
            else:
                txt = cell_text(r, c, cell)
            out.append(f'<td{attrs} style="{style}">{txt}</td>')
        out.append('</tr>')
    return '\n'.join(out)

# 印刷可能幅（A4縦 210mm − 左右0.25in）に合わせて縮小（狭い表は等倍まで）
PRINT_W = round((210 - 6.35 * 2) * 96 / 25.4)
PRINT_H = round((297 - 6.35 * 2) * 96 / 25.4)
zoom = round(min(PRINT_W / table_w, 1.0) * 0.99, 4)

# 大きすぎる案件ブロックは分割禁止を解除する。
# ・小さい案件は分割禁止のまま（＝途中で切れない）
# ・1ページの MAX_KEEP_FRACTION を超える案件は分割許可（＝大きな余白を作らない／はみ出さない）
MAX_KEEP_FRACTION = 0.65
def block_h(r0, r1):
    return sum(row_px(r) for r in range(r0, r1 + 1))
BLOCKS = [(r0, r1, keep and block_h(r0, r1) * zoom < PRINT_H * MAX_KEEP_FRACTION)
          for (r0, r1, keep) in BLOCKS]

css = f"""<meta charset="utf-8"><style>
@page {{ size: A4 portrait; margin: 0.25in; }}
* {{ box-sizing: border-box; }}
html, body {{ margin: 0; padding: 0; }}
body {{ font-family: {FONT}; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
.sheet {{ zoom: {zoom}; width: {table_w}px; margin: 0 auto; }}
table {{ border-collapse: collapse; table-layout: fixed; }}
tbody.keep {{ break-inside: avoid; page-break-inside: avoid; }}
td {{ padding: {CELL_PAD_Y}px 2px; line-height: {LINE_HEIGHT}; font-family: {FONT}; overflow: hidden; }}
</style>"""

parts = [css, f'<div class="sheet"><table>{COLGROUP}']
for (r0, r1, keep) in BLOCKS:
    cls = ' class="keep"' if (keep and KEEP_TOGETHER) else ''
    parts.append(f'<tbody{cls}>{render_rows(r0, r1)}</tbody>')
parts.append('</table></div>')

open(OUT, 'w', encoding='utf-8').write('\n'.join(parts))
print(f'sheet={ws.title!r} range=A1:{get_column_letter(MAXC)}{MAXR} '
      f'table_w={table_w} zoom={zoom} projects={len(proj_starts)} blocks={len(BLOCKS)}')
